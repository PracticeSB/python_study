import pyvisa
import numpy as np
import time
import math
import openpyxl #For making excel sheet

#######################################################
Loop_count = 1
Second_Loop_count = 0
Loop_count_cooling = 0
Loop_Total_time = 0
count = 1
Time_arr = []
Resistance_arr = []
Voltage_arr = []
Total_time_arr = []
Cycle = 0
Resistance = 0
Voltage = 0
Target_Voltage = 0
Total_time = 0
Cooling_time = 0
Cooling_resistance_arr = []
Cooling_time_arr = []
Cooling_loop_total_time = 0


wb = openpyxl.Workbook()
ws = wb.create_sheet("TEST")  # Name of the shneet
ws.cell(row=1, column=1, value='Time[Sec]')
ws.cell(row=1, column=2, value='Voltage[V]')
ws.cell(row=1, column=3, value='Resistance[Ohm]')
#ws.cell(row=1, column=4, value='Power[W]')
####################################################### Setting of the equipment
rm = pyvisa.ResourceManager()
Source_meter = rm.open_resource('TCPIP0::169.254.113.194::inst0::INSTR') #USB address
print(rm.list_opened_resources()) #For checking the address
print(Source_meter.query("*IDN?")) #For checking the response
Source_meter.write('*LANG SCPI')
Source_meter.write('*RST')
Source_meter.write('SENSe:FUNCtion "CURR"')
Source_meter.write('SENSe:CURRent:UNIT OHM')
Source_meter.write('SOURce:FUNCtion VOLT')
Source_meter.write('TRACe:CLear "defbuffer1"') #Clear the buffer
Source_meter.write('TRACe:CLear "defbuffer2"') #Clear the buffer
#######################################################

#######################################################
Initial_Voltage = 1
Source_meter.write('SOURce:VOLT '+str(Initial_Voltage))
Source_meter.write('SOURce:VOLT:ILIM 0.1') #Limitation of the current
Source_meter.write('OUTPut ON') #Turn the output on
Source_meter.write(':INIT') #Initiate readings
Source_meter.write(':WAI') #Wait for the measurement to finish
Source_meter.write('TRACe:TRIGger "defbuffer2"')
for i in range(1,2):
    Initial_resistance = Source_meter.query('TRACe:DATA? ' +str(i)+',' +str(i)+',"defbuffer2", READ')
Input_Power = (0.125)*9
Target_Voltage = (math.sqrt(float(Input_Power)*float(Initial_resistance))) #Power = V^2/R  --> V = sqrt(Power*R)
print(Target_Voltage)
if Target_Voltage > 70:
    print('ERROR')
    print(Target_Voltage)
    Source_meter.write('OUTPUT OFF')
    exit()
#######################################################

####################################################### Loop start
while Cycle < 1:
    while True:
        Source_meter.write('OUTPut ON') #Turn the output on
        First_start = time.time()  # For Loop time
        Loop_start_time = time.time()
        if Target_Voltage > 21:
            Source_meter.write('SOURce:VOLT:ILIM 0.1') #Limitation of the current
            if Target_Voltage > 60:
                Source_meter.write('OUTPUT OFF')
        if Target_Voltage <= 21:
            Source_meter.write('SOURce:VOLT:ILIM 1') #Limitation of the current
        Source_meter.write('SOURce:VOLT '+str(Target_Voltage)) #Target Voltage와 연동시키기
        Source_meter.write(':READ?')
        Measured = Source_meter.read()
        Measured_voltage = math.sqrt(float(Input_Power)*float(Measured))
        v = Measured_voltage
        Voltage_arr.append(Measured_voltage)
        Resistance_arr.append(float(Measured))
        r = Measured
        Voltage = (math.sqrt(float(Input_Power)*float(Measured)))
        Target_Voltage = Voltage
        count = count + 1
        First_end = time.time() #For Loop time
        Loop_end_time = time.time()
        Loop_Interval = Loop_end_time - Loop_start_time
        Loop_Total_time = Loop_Total_time + Loop_Interval
        Interval = First_end - First_start
        Total_time = Total_time + Interval
        ws.cell(row=Loop_count+Second_Loop_count, column=1, value=Total_time)
        ws.cell(row=Loop_count+Second_Loop_count, column=2, value=Voltage)
        ws.cell(row=Loop_count+Second_Loop_count, column=3, value=float(r))
        Loop_count = Loop_count + 1
        Total_time_arr.append(Total_time)
        if Loop_Total_time > 20: #Voltage를 가해주는 시간
            Source_meter.write('SOURce:VOLT '+str(0.01))
            while Cooling_loop_total_time < 80:
                Cooling_start = time.time()
                Cooling_Loop_start_time = time.time()
                r = Source_meter.query(':MEAS?')
                Voltage_arr.append(float(0))
                Cooling_end = time.time()
                Cooling_Loop_end_time = time.time()
                Cooling_Interval = Cooling_end - Cooling_start
                Cooling_time = Cooling_time + Cooling_Interval
                Cooling_loop_Interval = Cooling_Loop_end_time - Cooling_Loop_start_time
                Total_time = Total_time + Cooling_Interval
                Total_time_arr.append(Total_time)
                Cooling_loop_total_time = Cooling_loop_total_time + Cooling_loop_Interval
                ws.cell(row=Loop_count + Second_Loop_count, column=1, value=Total_time)
                ws.cell(row=Loop_count + Second_Loop_count, column=2, value=float(0))
                ws.cell(row=Loop_count + Second_Loop_count, column=3, value=float(r))
                Second_Loop_count = Second_Loop_count + 1
            Loop_Total_time = 0
            Cooling_loop_total_time = 0
            Cycle = Cycle + 1
            print(Cycle)
            break
Source_meter.write('OUTPUT OFF')
####################################################### Loop end
wb.save(r'D:\Dropbox\노승범\실험\TOF\Laser\# of the parallel\10parallel\20230220\sample2\Source_Experiment_0.125W_30g_9cm.xlsx')
print('Data is saved')
