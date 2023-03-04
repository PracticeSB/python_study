import threading
import time
import keyboard
import numpy as np
import math
import pyvisa
import matplotlib.pyplot as plt
import matplotlib.animation as animation
import openpyxl #For making excel sheet

####################################################### Variable data
Sample_length = 0
Sample_watt = 0
####################################################### Don't touch
First_start = 0
First_end = 0
Measured = 0
Measured_voltage = 0
Cycle = 0
Resistance = 0
Voltage = 0
Target_Voltage = 0
Total_time = 0
Delta_r_percentage = 0
error = 0
Time_arr = []
Resistance_arr = []
Voltage_arr = []
Total_time_arr = []
Loop_count = 1
Target_Voltage = 1
i_control = 0
Interval = 0
Kp = 20
Ki = 10
flag = 0
Set_point = 1
Value = 400
####################################################### For excel
wb = openpyxl.Workbook()
ws = wb.create_sheet("TEST")  # Name of the sheet
ws.cell(row=1, column=1, value='Time[Sec]')
ws.cell(row=1, column=2, value='Voltage[V]')
ws.cell(row=1, column=3, value='Resistance[Ohm]')
####################################################### Create figure for plotting
fig = plt.figure()
ax = fig.add_subplot(3, 1, 1)
ax1 = fig.add_subplot(3, 1, 2)
ax2 = fig.add_subplot(3, 1, 3)
x = []
y = []
y1 = []
y2 = []

#######################################################
def map(x,input_min,input_max,output_min,output_max):
    return (x-input_min)*(output_max-output_min)/(input_max-input_min)+output_min
#######################################################
def data_query():
    global Resistance
    global Interval
    global Total_time
    global Target_Voltage
    global Delta_r_percentage
    global Set_point
    global error
    global i_control
    global p_control
    global Loop_count
    global Value
    global flag
    while Total_time < 60:
        if keyboard.is_pressed("a"):
            Set_point = 10
        if keyboard.is_pressed("b"):
            Set_point = 20
        if keyboard.is_pressed("c"):
            Set_point = 25
        if keyboard.is_pressed("d"):
            Set_point = 30
        if keyboard.is_pressed("e"):
            Set_point = 40
        if keyboard.is_pressed("f"):
            Set_point = 50
        First_start = time.time()
        Source_meter.write('SOURce:VOLT:ILIM 1') #Limitation of the current
        Source_meter.write('OUTPut ON') #Turn the output on
        Source_meter.write('SOURce:VOLT '+str(Target_Voltage))
        Source_meter.write(':READ?')
        r = Source_meter.read()
        Resistance = float(r)
        Delta_r_percentage = -((Resistance-float(Initial_resistance))/float(Initial_resistance))*100
        error = Set_point-Delta_r_percentage
        print("Error:",error)
        p_control = Kp*error
        i_control = i_control+Ki*error*Interval
        print("Interval:", Interval)
        pi = p_control+i_control #Output data
        print("pi:", pi)
        if pi > Value:
            pi = Value
        if error < 0:
            pi = 1
        Target_Voltage = map(pi,0,Value,0,20) #pi control의 값을 voltage로 치환하여서 대입함.
        print("Target Voltage:",Target_Voltage)
        First_end = time.time()
        Interval = First_end - First_start
        Total_time = Total_time + Interval
        ws.cell(row=Loop_count, column=1, value=Total_time)
        ws.cell(row=Loop_count, column=2, value=Target_Voltage)
        ws.cell(row=Loop_count, column=3, value=float(r))
        ws.cell(row=Loop_count, column=4, value=Delta_r_percentage)
        Loop_count = Loop_count + 1
    Source_meter.write('OUTPUT OFF')
    wb.save(r'D:\Dropbox\노승범\실험\TOF\Laser\PI_control\20230220\%d_%0.2f_9.5cm_c.xlsx' %(Kp,Ki))
    print('Data is saved')
#######################################################
def animate(i, x, y,y1,y2):
    x.append(Total_time)
    y.append(Resistance)
    y1.append(Delta_r_percentage)
    y2.append(error)
    if len(x) > 50:
        x = x[-50:]
        y = y[-50:]
        y1 = y1[-50:]
        y2 = y2[-50:]
    ax.clear()
    ax1.clear()
    ax2.clear()
    ax.plot(x, y, 'r-', label='Resistance')
    ax.set_ylabel('Resistance(ohms)')
    ax1.plot(x, y1, 'b-')
    ax1.set_ylabel(r'$\Delta$Resistance/R0')
    ax2.plot(x, y2, 'y-')
    ax2.set_ylabel('Error')
    ax.legend(loc=(0.7,0.9), frameon=False, shadow=False)
    fig.suptitle('Real-time Sensing Data', fontsize = 20)
    plt.xlabel('Time(Sec)')
####################################################### Setting of the equipment
rm = pyvisa.ResourceManager()
Source_meter = rm.open_resource('TCPIP0::169.254.113.194::inst0::INSTR') #LAN communication address
print(rm.list_opened_resources()) #For checking the address
print(Source_meter.query("*IDN?")) #For checking the response
Source_meter.write('*LANG SCPI')
Source_meter.write('*RST')
Source_meter.write('SENSe:FUNCtion "CURR"')
Source_meter.write('SENSe:CURRent:UNIT OHM')
Source_meter.write('SOURce:FUNCtion VOLT')
Source_meter.write('TRACe:CLear "defbuffer1"') #Clear the buffer
Source_meter.write('TRACe:CLear "defbuffer2"') #Clear the buffer
#######################################################

####################################################### Checking the initial resistance of fiber actuator
Initial_Voltage = 1
Source_meter.write('SOURce:VOLT '+str(Initial_Voltage))
Source_meter.write('SOURce:VOLT:ILIM 0.1') #Limitation of the current
Source_meter.write('OUTPut ON') #Turn the output on
Source_meter.write(':INIT') #Initiate readings
Source_meter.write(':WAI') #Wait for the measurement to finish
Source_meter.write('TRACe:TRIGger "defbuffer2"')
for i in range(1,2):
    Initial_resistance = Source_meter.query('TRACe:DATA? ' +str(i)+',' +str(i)+',"defbuffer2", READ')
####################################################### Over voltage regulation
if __name__ == "__main__":
    thread1 = threading.Thread(target=data_query)
    thread1.start()
    thread1.daemon
    ani = animation.FuncAnimation(fig, animate, fargs=(x, y, y1, y2), interval=Interval)
    plt.show()
