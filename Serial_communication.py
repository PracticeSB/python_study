import serial
import threading
import time
import matplotlib.pyplot as plt
import matplotlib.animation as animation
import openpyxl
from queue import Queue

#RS485는 half duplex 방식의 통신이므로 송신이 끝나고 그 다음에 바로 수신이 되어야함
q = Queue(maxsize=0)
flag = True

#Buffer for memorizing sensor data
Data = []
Total_time = 0
Response = 0
Interval = 0
Loop_count = 0

#For making excel sheet
wb = openpyxl.Workbook()
ws = wb.create_sheet("TEST")  # Name of the sheet
ws.cell(row=1, column=1, value='Time[Sec]')
ws.cell(row=1, column=2, value='Displacement[mm]')

# Create figure for plotting
fig = plt.figure()
ax = fig.add_subplot(1, 1, 1)
x = []
y = []

class Laser():
    def __init__(self,com,baudrate):
        self.mySerial = serial.Serial(com,baudrate,timeout = 0,bytesize=serial.EIGHTBITS)
    def SendData(self,q):
        global flag
        while True:
            if flag == True:
                packet = bytes(bytearray([0x02,0x43,0xB0,0x01,0x03,0xF2]))
                self.mySerial.write(packet)
                #print('Send :' + str(packet))
                flag = False
                q.put(flag)
    def Setup(self,dataPacket):
        packet = bytes(bytearray(dataPacket))
        self.mySerial.write(packet)
        print('Setup :' + str(packet))
    def Setpoint(self,dataPacket):
        packet = bytes(bytearray(dataPacket))
        self.mySerial.write(packet)
        print('Setpoint :' + str(packet))
    def RecvData(self,q):
        global flag
        global Total_time
        global Response
        global Interval
        global Loop_count
        while Total_time < 30:
            start_time = time.time()
            flag = q.get()
            if flag == False:
                if self.mySerial.readable():
                    A = bytearray(self.mySerial.read(6))
                    if len(A) < 6:
                        B = self.mySerial.read(size=6-len(A))
                        A.extend(B)
                        #print('Additional' + str(A) , len(A))
                    #### 만약에 6보다 긴 경우가 많이 나오면 수정해야할 부분
                    # if len(A) > 6:
                    #     for i in range(len(A) - 1):
                    #         if A[i] == bytes(0x02):
                    #             for count in range(4):
                    #                 Data.append(A[i+count])
                    #                 print('Over 6' + Data)
                    if A[2] >= 127:
                        Response = -((255 - (A[2])) * 256 + (255 - A[3])) * 0.01
                        B_calibration = (Response * 1.18 - 0.16534)
                    else:
                        Response = ((A[2]) * 256 + (A[3])) * 0.01
                        B_calibration = (Response * 1.18 + 0.16534)
                    #print(A,len(A))
                    #print('Response: ', Response,'\n')
                    #time.sleep(0.05)
                    end_time = time.time()
                    Interval = end_time - start_time
                    #print('Interval', Interval, '\n')
                    Total_time = Total_time + Interval
                    print(Total_time)
                    Loop_count = Loop_count + 1
                    ws.cell(row=Loop_count + 2, column=1, value=Total_time)
                    ws.cell(row=Loop_count + 2, column=2, value=Response)
                    ws.cell(row=Loop_count + 2, column=3, value=B_calibration)
                    flag = True
        self.mySerial.write(bytes(bytearray([0x02, 0x43, 0xA1, 0x01, 0x03, 0xE3])))  # release set point
        time.sleep(1)
        wb.save(r'C:\Users\user\PycharmProjects\Data\1.csv')
        print('Data is saved!')
def animate(i, x, y):
    x.append(Total_time)
    y.append(Response)
    if len(x) > 50:
         x = x[-50:]
         y = y[-50:]
    ax.clear()
    ax.plot(x, y, 'r-', label='Distance')
    ax.legend(loc=(0.7,0.9), frameon=False, shadow=False)
    plt.title('Real-time Sensing Data', fontsize = 20)
    plt.xlabel('Time(Sec)')
    plt.ylabel('Displacement(mm)')



A = Laser("COM5", 115200)
A.Setup([0x02,0x43,0xA1,0x01,0x03,0xE3])
time.sleep(1)
A.Setup([0x02,0x43,0xA1,0x00,0x03,0xE2])
time.sleep(1)

if __name__=='__main__':
    t1 = threading.Thread(target=A.SendData,args=(q,))
    t2 = threading.Thread(target=A.RecvData, args=(q,))
    t1.start()
    t2.start()
    ani = animation.FuncAnimation(fig, animate, fargs=(x, y), interval=Interval)
    plt.show()









