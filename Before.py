import threading
import serial
import time
import numpy as np
import openpyxl #For making excel sheet
import matplotlib.pyplot as plt
import matplotlib.animation as animation

Data = []
Median_Value  = []
Window_list = []
C = 0
B_calibration = 0
Total_time = 0
Loop_count = 0

baud = 115200
time_value = 0.3

# Create figure for plotting
fig = plt.figure()
ax = fig.add_subplot(1, 1, 1)
x = []
y = []

wb = openpyxl.Workbook()
ws = wb.create_sheet("TEST")  # Name of the sheet
ws.cell(row=1, column=1, value='Time[Sec]')
ws.cell(row=1, column=2, value='Displacement[mm]')

ser = serial.Serial(port='COM5',baudrate=baud,timeout= time_value)
ser.write(bytes(bytearray([0x02,0x43,0xA1,0x01,0x03,0xE3]))) #release set point
time.sleep(1)
ser.write(bytes(bytearray([0x02,0x43,0xA1,0x00,0x03,0xE2]))) #Set point
time.sleep(1)

def read():
    global Data
    global Total_time
    global Loop_count
    global C
    global B_calibration
    while Total_time <= 600:
        if ser.readable():
            Start_time = time.time()
            A = bytearray(ser.readline())
            _list = A
            print(_list,len(_list))
            for i in range(len(_list)-1):
                if ((_list[i]) == 2) and ((_list[i+1]) == 6):
                    for count in range(4):
                        Data.append((_list[i+count]))
                        try:
                            if Data[2] >= 127:
                                Response = -((255-(Data[2]))*256 + (255-Data[3]))*0.01
                                B_calibration = (Response*1.18-0.16534)
                            else:
                                Response = ((Data[2])*256 + (Data[3]))*0.01
                                B_calibration = (Response*1.18+0.16534)
                            Window_list.append(B_calibration)
                            if len(Window_list) > 5:
                                Window_list.pop(0)
                            C = np.median(Window_list)
                            print(C)
                            Median_Value.append(C)
                        except IndexError:
                            pass
                Data= []
            End_time = time.time()
            Time_Interval = End_time - Start_time
            #print(Time_Interval)
            Total_time = Total_time + Time_Interval
            Loop_count = Loop_count + 1
            #print(Total_time)
            ws.cell(row=Loop_count + 2, column=1, value=Total_time)
            ws.cell(row=Loop_count + 2, column=2, value=C)
            ws.cell(row=Loop_count + 2, column=3, value=B_calibration)
    ser.write(bytes(bytearray([0x02,0x43,0xA1,0x01,0x03,0xE3]))) #release set point
    time.sleep(1)
    wb.save(r'D:\Dropbox\노승범\실험\TOF\Laser\Sensing\20221109\Laser_Experiment_0.2W_80g_3.5cm_1.csv')
    print('Data is saved!')

def data_query():
    while True:
        ser.write(bytes(bytearray([0x02,0x43,0xB0,0x01,0x03,0xF2])))
        time.sleep(time_value)

def animate(i, x, y):
    x.append(Total_time)
    y.append(C)
    if len(x) > 50:
        x = x[-50:]
        y = y[-50:]
    ax.clear()
    ax.plot(x, y, 'r-', label='Distance')
    ax.legend(loc=(0.7,0.9), frameon=False, shadow=False)
    plt.title('Real-time Sensing Data', fontsize = 20)
    plt.xlabel('Time(Sec)')
    plt.ylabel('Displacement(mm)')

if __name__ == "__main__":
    thread1 = threading.Thread(target= data_query)
    thread2 = threading.Thread(target=read)
    thread1.start()
    thread1.daemon
    thread2.start()
    thread2.daemon
    ani = animation.FuncAnimation(fig, animate, fargs=(x, y), interval=time_value)
    plt.show()