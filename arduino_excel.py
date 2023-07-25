import openpyxl
import time
import keyboard
import struct
import serial
import serial.tools.list_ports as sp

d = time.strftime('%Y-%m-%d-%H-%M', time.localtime((time.time())))
print("Date:",d)

wb = openpyxl.Workbook('')
ws = wb.create_sheet('Sheet1')

ws.cell(row=1, column=1, value = 'Time(Sec)')
ws.cell(row=1, column=2, value = 'current_1')
ws.cell(row=1, column=3, value = 'voltage_1')
ws.cell(row=1, column=4, value = 'current_2')
ws.cell(row=1, column=5, value = 'voltage_2')
ws.cell(row=1, column=6, value = 'current_3')
ws.cell(row=1, column=7, value = 'voltage_3')
ws.cell(row=1, column=8, value = 'current_4')
ws.cell(row=1, column=9, value = 'votlage_4')

#comports returns a list containing ListPortInfo objects
list = sp.comports()
print("Number of the ports:",len(list))

for port, desc, hwid in sorted(list):
        print("{}: {} [{}]".format(port, desc, hwid))

py_serial = serial.Serial(port = "COM8", baudrate = 115200,)

A = []
Loop_count = 0
sensor_number = 0


c = input('s를 입력해주세요')
if c == 's':
    c = c.encode('utf-8')
    py_serial.write(c)
    print('아두이노에 시리얼을 보냈습니다')

while (True):
    if keyboard.is_pressed ("a"):
        break
    if py_serial.readable():
        response = py_serial.readline()
        A = response.split() #split 인자가 없다면 공백에서 split
        print(A)
        try:
            ws.cell(row=Loop_count+2, column=1, value=float(A[0]))
            ws.cell(row=Loop_count+2, column=2*sensor_number+2, value=float(A[1])*0.001) #current mA단위라서 단위환산
            ws.cell(row=Loop_count+2, column=2*sensor_number+3, value=float(A[2]))
            sensor_number = sensor_number + 1
            if sensor_number > 3:
                Loop_count = Loop_count + 1
                sensor_number = 0
        except:
            print('예외가 발생하였습니다')
            pass

py_serial.close() #stop serial communication
wb.save(r'D:\Dropbox\노승범\실험\Application\Experiment\Experiment '+ d +'.xlsx')
print('저장이 완료되었습니다')
